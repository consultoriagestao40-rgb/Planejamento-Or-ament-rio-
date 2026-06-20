import openpyxl

def main():
    try:
        wb = openpyxl.load_workbook('/Users/cristianosilva/BudgetHub/planilha.xlsx', read_only=True)
        print("Sheets in planilha.xlsx:")
        for sheet in wb.sheetnames:
            print(f"  Sheet: {sheet}")
            ws = wb[sheet]
            # Print first 10 rows
            for idx, r in enumerate(ws.iter_rows(values_only=True)):
                if idx >= 10:
                    break
                print(f"    Row {idx}: {r[:10]}")
    except Exception as e:
        print("Error reading planilha.xlsx:", e)

if __name__ == '__main__':
    main()
